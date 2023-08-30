import openpyxl




# 将excel转换成二维数组
def excel_to_2d_array(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

# 将excel的第一行转换成一维数组
def excel_first_row_to_array(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    return first_row

# 查找数据中值的索引
def find_index_of_value(array, value):
    try:
        index = array.index(value)
        return index
    except ValueError:
        return None

# 获得二维数组中的值
def get_value_2d_array(array_2d, row_index, col_index):
    if row_index < len(array_2d) and col_index < len(array_2d[0]):
        return array_2d[row_index][col_index]
    else:
        return None

if __name__ == '__main__':
    # 替换为你的Excel文件路径和工作表名称
    data_file_path = 'data.xlsx'
    data_sheet_name = 'Sheet1'
    v_sheet_name = 'Sheet2'
    a_sheet_name = 'Sheet3'

    # 将excel 转换成数组
    data_2d_array = excel_to_2d_array(data_file_path, data_sheet_name)
    v_array = excel_first_row_to_array(data_file_path, v_sheet_name)
    a_array = excel_first_row_to_array(data_file_path, a_sheet_name)

    v = find_index_of_value(v_array, 5)
    a = find_index_of_value(a_array, -8)
    print(v, a)

    value = get_value_2d_array(data_2d_array, v, a)
    print(value)
